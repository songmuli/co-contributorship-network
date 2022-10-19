import xml.etree.cElementTree as ET
import os
import os.path
import xlwt
import openpyxl

path = 'D:/2022bishe/pone/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('pone', index=0)
ws["A1"] = "摘要"
i=2

for xmlFile in files:
    tree = ET.parse(os.path.join(path, xmlFile))
    root = tree.getroot()
    abstracts = root.find("./front/article-meta/abstract")
    # print(abstracts)
    # print(xmlFile)
    abstracts_list = []
    try:
       for str in abstracts.itertext():
         abstracts_list.append(str.strip())
       abstract  = ' '.join(abstracts_list)
    except AttributeError:
        abstract = 'None'
    print(abstract)
    # with open('D:/2022bishe/abstract_table/pone_abstract.txt', 'a+', encoding="utf-8") as f:
    #  f.write(abstract)
    #  f.write('\n\n')
    #  f.close()
    ws.cell(row=i, column=1, value=abstract)
    i += 1
wb.save('D:/2022bishe/abstract_table/pone_abstact.xlsx')
#    i += 1
# print(i)
# tree = ET.parse('D:/2022bishe/pbio/file@id=10.1371%2Fjournal.pbio.1002332&type=manuscript')
# root = tree.getroot()
# abstracts = root.find("./front/article-meta/abstract")
# # print(abstract)
# abstract = []
# for str in abstracts.itertext():
#   abstract.append(str)
#    # i += 1
# abstract_str = ''.join(abstract)
# print(abstract_str)
