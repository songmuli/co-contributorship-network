import xml.etree.cElementTree as ET
import os
import os.path
import openpyxl

path = 'D:/2022bishe/xml/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('pone', index=0)
ws["A1"] = "文章名称"
i=2
for xmlFile in files:
   tree = ET.parse(os.path.join(path, xmlFile))
   root = tree.getroot()
   articles = root.find("./front/article-meta/title-group/article-title")
   # print(articles)
   articles_title = []
   for str in articles.itertext():
      articles_title.append(str)
   article_title = ''.join(articles_title)
   # print(article_title)
   ws.cell(row=i, column=1, value=article_title)
   i += 1
wb.save('D:/2022bishe/pone_article-title.xlsx')