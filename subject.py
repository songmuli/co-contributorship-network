import xml.etree.cElementTree as ET
import os.path
import openpyxl

path = 'D:/2022bishe/ppat/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('ppat', index=0)
ws["A1"] = "DOI"
ws["B1"] = "学科信息"
i = 2

for xmlFile in files:
  # print(xmlFile)
  tree = ET.parse(os.path.join(path, xmlFile))
  root = tree.getroot()
  subgroups = root.findall('./front/article-meta/article-categories/subj-group')
  sub_information = []
  for subgroup in subgroups:
    if subgroup.attrib['subj-group-type'] == 'Discipline-v3':
        subject=subgroups
        subjects = []
        for str in subgroup.itertext():
          if str != '\n':
            subjects.append(str.strip())
        subject = ' -- '.join(subjects)
        sub_information.append(subject)
  Doi = ''
  dois = root.findall("./front/article-meta/article-id")
  for doi in dois:
      if doi.attrib['pub-id-type'] == 'doi':
          Doi = doi.text
  ws.cell(row=i, column=1, value=Doi)
  ws.cell(row=i, column=2, value=' ; '.join(sub_information))
  i += 1
wb.save('D:/2022bishe/subject_table/ppat.xlsx')
  # print(' ; '.join(sub_information))