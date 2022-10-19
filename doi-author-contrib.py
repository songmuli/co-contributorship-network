import xml.etree.cElementTree as ET
import os.path
import openpyxl

path = 'D:/2022bishe/ppat/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('ppat', index=0)
ws["A1"] = "DOI"
ws["B1"] = "作者"
ws["C1"] = "所做贡献"
i = 2

for xmlFile in files:
  tree = ET.parse(os.path.join(path, xmlFile))
  root = tree.getroot()
  year = root.find("./front/article-meta/pub-date/year").text
  if year != "2016":
    contribs = root.findall("./front/article-meta/contrib-group/contrib")
    for contrib in contribs:
    # print(contrib.attrib)
       if contrib.attrib['contrib-type'] == "author":
         try:
            surnames = contrib.find("./name/surname").text
            given_names = contrib.find("./name/given-names").text
            names = given_names + ' ' + surnames
          # print(names)
         except:
            names = contrib.find("./collab").text

         contributions = []
         roles = contrib.findall("./role")
         for role in roles:
             contribution = role.text
             contributions.append(contribution)
         contribu = " ; ".join(str(j) for j in contributions)

         Doi = ''
         dois = root.findall("./front/article-meta/article-id")
         for doi in dois:
             if doi.attrib['pub-id-type'] == 'doi':
                 Doi = doi.text

         if (len(contributions) == 0):
             print("贡献数据不规范")
         else:
             ws.cell(row=i, column=1, value=Doi)
             ws.cell(row=i, column=2, value=names)
             ws.cell(row=i, column=3, value=contribu)
             i += 1
wb.save('D:/2022bishe/doi-author-contrib_table/ppat.xlsx')
             # print(Doi,  names, contribu )