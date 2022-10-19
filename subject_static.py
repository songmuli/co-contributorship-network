import pymysql
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet('pone', index=0)
ws["A1"] = "first_subject"
ws["B1"] = "second_subject"
ws["C1"] = "Doi"
i = 2

db = pymysql.connect(host='localhost', user='root', password='123456', db='bishe')
print("数据库连接成功！")
cursor = db.cursor()



subject_bio = []
subject_med = []
subject_res = []
subject_eco = []
subject_phy = []
subject_eng = []
subject_ear = []
subject_peo = []
subject_com = []
subject_soc = []
subject_sci = []
subject_vig = []

# cursor.execute(" SELECT DISTINCT second_subject FROM subject_quchong WHERE first_subject='Biology and life sciences' ")
# result_bio_1 = cursor.fetchall()
# for res in result_bio_1:
#     subject_bio.append(res[0])
#
# for bio in subject_bio:
#   doi_bio = []
#   if bio != None:
#     cursor.execute('SELECT doi FROM subject_quchong where (first_subject = %s) & (second_subject = %s)', ('Biology and life sciences', bio))
#     result_bio_2 = cursor.fetchall()
#     for res in result_bio_2:
#       doi_bio.append(res[0])
#   else:
#       cursor.execute('SELECT doi FROM subject_quchong where (first_subject = %s) & (second_subject is null)', ('Biology and life sciences'))
#       result_bio_3 = cursor.fetchall()
#       for res in result_bio_3:
#           doi_bio.append(res[0])
#   for doi in doi_bio:
#         # print('Biology and life sciences', bio, doi)
#     ws.cell(row=i, column=1, value='Biology and life sciences')
#     ws.cell(row=i, column=2, value=bio)
#     ws.cell(row=i, column=3, value=doi)
#     i += 1

cursor.execute(" SELECT DISTINCT second_subject FROM subject_quchong WHERE first_subject='Vigilance (psychology)' ")
result_1 = cursor.fetchall()
for res in result_1:
    subject_vig.append(res[0])

for vig in subject_vig:
  doi_vig = []
  if vig != None:
    cursor.execute('SELECT doi FROM subject_quchong where (first_subject = %s) & (second_subject = %s)', ('Vigilance (psychology)', vig))
    result_2 = cursor.fetchall()
    for res in result_2:
      doi_vig.append(res[0])
  else:
      cursor.execute('SELECT doi FROM subject_quchong where (first_subject = %s) & (second_subject is null)', ('Vigilance (psychology)'))
      result_3 = cursor.fetchall()
      for res in result_3:
          doi_vig.append(res[0])
  for doi in doi_vig:
        # print('Biology and life sciences', bio, doi)
    ws.cell(row=i, column=1, value='Vigilance (psychology)')
    ws.cell(row=i, column=2, value=vig)
    ws.cell(row=i, column=3, value=doi)
    i += 1

wb.save('D:/2022bishe/total/subject/subject_vig.xlsx')