import xml.etree.cElementTree as ET
import os
import os.path
import openpyxl

# path = 'D:/2022bishe/pbio/file@id=10.1371%2Fjournal.pbio.1002332&type=manuscript'
path = 'D:/2022/xml/pbio/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('pone', index=0)
ws["A1"] = "doi"
ws["B1"] = "author_name"
ws["C1"] = "author_corresp"
ws["D1"] = "author_aff"
i = 2

for xmlFile in files:
  tree = ET.parse(os.path.join(path, xmlFile))
  root = tree.getroot()

  Doi = ''
  dois = root.findall("./front/article-meta/article-id")
  for doi in dois:
    if doi.attrib['pub-id-type'] == 'doi':
      Doi = doi.text

  contribs = root.findall("./front/article-meta/contrib-group/contrib")
  affs = root.findall("./front/article-meta/aff")
# for aff in affs:
#     addr = aff.find("./addr-line").text
#     print(addr)
  for contrib in contribs:
    if contrib.attrib['contrib-type'] == 'author':
     try:
       surnames = contrib.find("./name/surname").text
       given_names = contrib.find("./name/given-names").text
       names = given_names + '  ' + surnames
      # print(names)
     except:
         names = contrib.find("./collab").text
     addr_line = []
     addr = ''
     corresp = ''
     xrefs = contrib.findall("./xref")
     for xref in xrefs:
        if xref.attrib["ref-type"] == 'corresp':
            corresp = ' Y '

        if xref.attrib["ref-type"] == 'aff':
          rid = xref.attrib['rid']
          for aff in affs:
           if aff.attrib["id"] == rid:
            addr_line.append(aff.find("./addr-line").text)
          addr = ' ; '.join(str(addrline) for addrline in addr_line)

     ws.cell(row=i, column=1, value=Doi)
     ws.cell(row=i, column=2, value=names)
     ws.cell(row=i, column=3, value=corresp)
     ws.cell(row=i, column=4, value=addr)
     i += 1
  print(xmlFile)
wb.save('D:/2022/data/author-information/pbio_author-information.xlsx')
  #      information = names + " : " + addr
  #      print(information)

