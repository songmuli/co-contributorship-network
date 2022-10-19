import xml.etree.cElementTree as ET
import os.path
import openpyxl

path = 'D:/2022bishe/xml/pone-1/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('pone', index=0)
ws["A1"] = "DOI"
ws["B1"] = "first_subject"
ws["C1"] = "second_subject"
i = 2

for xmlFile in files:
  print(xmlFile)
  tree = ET.parse(os.path.join(path, xmlFile))
  root = tree.getroot()

  Doi = ''
  dois = root.findall("./front/article-meta/article-id")
  for doi in dois:
    if doi.attrib['pub-id-type'] == 'doi':
      Doi = doi.text

  subgroups = root.findall('./front/article-meta/article-categories/subj-group')
  first_subject = []
  second_subject = []
  for subgroup in subgroups:
    if subgroup.attrib['subj-group-type'] == 'Discipline-v3':
        f_subject = subgroup.find("./subject").text
        first_subject.append(f_subject)
        try:
          s_subject = subgroup.find("./subj-group/subject").text
          second_subject.append(s_subject)
        except:
          s_subject = ''
          second_subject.append(s_subject)
  for f,s in zip(first_subject, second_subject):
    print(Doi, f, s)
    ws.cell(row=i, column=1, value=Doi)
    ws.cell(row=i, column=2, value=f)
    ws.cell(row=i, column=3, value=s)
    i += 1
wb.save('D:/2022bishe/total/subject_pone-2.xlsx')
