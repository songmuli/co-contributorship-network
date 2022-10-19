import xml.etree.cElementTree as ET
import os.path
import openpyxl


path = 'D:/2022bishe/xml/pone-1/'
files = os.listdir(path)
wb = openpyxl.Workbook()
ws = wb.create_sheet('pbio', index=0)
ws["A1"] = "DOI"
ws["B1"] = "contrib_type"
ws["C1"] = "author_name"
i = 2
contributions = ["Conceptualization", "Formal analysis", "Investigation", "Methodology", "Visualization", "Writing – original draft", "Writing – review & editing", "Resources", "Supervision", "Funding acquisition", "Data curation", "Project administration", "Validation", "Software"]
for xmlFile in files:
  print(xmlFile)
  tree = ET.parse(os.path.join(path, xmlFile))
  root = tree.getroot()
  year = root.find("./front/article-meta/pub-date/year").text
  if year != "2016":
    contribs = root.findall("./front/article-meta/contrib-group/contrib")
    # print(xmlFile)
    Concept = []
    Analysis = []
    Investigation = []
    Method = []
    Visualization = []
    Draft = []
    Review = []
    Resources = []
    Supervision = []
    Funding = []
    Data = []
    Project = []
    Validation = []
    Software = []
    Others = []
    for contrib in contribs:
    # print(contrib.attrib)
      if contrib.attrib['contrib-type'] == "author":
        roles = contrib.findall("role")
        for role in roles:
            if role.text == "Conceptualization":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Concept.append(author_name)
            elif role.text == "Formal analysis":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Analysis.append(author_name)
            elif role.text == "Investigation":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Investigation.append(author_name)
            elif role.text == "Methodology":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Method.append(author_name)
            elif role.text == "Visualization":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Visualization.append(author_name)
            elif role.text == "Writing – original draft":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Draft.append(author_name)
            elif role.text == "Writing – review & editing":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Review.append(author_name)
            elif role.text == "Resources":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Resources.append(author_name)
            elif role.text == "Supervision":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Supervision.append(author_name)
            elif role.text == "Funding acquisition":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Funding.append(author_name)
            elif role.text == "Data curation":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Data.append(author_name)
            elif role.text == "Project administration":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Project.append(author_name)
            elif role.text == "Validation":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Validation.append(author_name)
            elif role.text == "Software":
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Software.append(author_name)
            else:
                author_name = contrib.find("./name/given-names").text + ' ' + contrib.find("./name/surname").text
                Others.append(author_name)
    # print(Draft)
    Doi = ''
    dois = root.findall("./front/article-meta/article-id")
    for doi in dois:
        if doi.attrib['pub-id-type'] == 'doi':
            Doi = doi.text
    if (len(Concept) == 0) & (len(Analysis) == 0) & (len(Investigation) == 0) & (len(Method) == 0) & (len(Visualization) == 0) & (len(Draft) == 0) & (len(Review) == 0) & (len(Resources) == 0) & (len(Supervision) == 0) & (len(Funding) == 0) & (len(Data) == 0) & (len(Project) == 0) & (len(Validation) == 0) & (len(Software) == 0):
        print("列表为空")
    else:
        # contributions ='Conceptualization: ' + " , ".join(str(contribution) for contribution in Concept) +' ; ' + 'Formal analysis: ' + " , ".join(str(analysis) for analysis in Analysis) +' ; '  + 'Investigation: ' + " , ".join(str(investigation) for investigation in Investigation) +' ; '  + 'Methodology: ' + " , ".join(str(method) for method in Method) +' ; '  + 'Visualization: ' + " , ".join(str(visualization) for visualization in Visualization) +' ; '  + 'Writing – original draft: ' + " , ".join(str(draft) for draft in Draft) +' ; '  + 'Writing – review & editing: ' + " , ".join(str(review) for review in Review) +' ; '  + 'Resources: ' + " , ".join(str(resources) for resources in Resources) +' ; '  + 'Supervision: ' + " , ".join(str(supervision) for supervision in Supervision) +' ; '  + 'Funding acquisition: ' + " , ".join(str(fund) for fund in Funding) +' ; '  + 'Data curation: ' + " , ".join(str(data) for data in Data) +' ; '  + 'Project administration: ' + " , ".join(str(project) for project in Project) +' ; '  + 'Validation: ' + " , ".join(str(valid) for valid in Validation) +' ; '  + 'Software: ' + " , ".join(str(software) for software in Software) +' ; '
     # print(contributions)
        for j in contributions:
           if (j == "Conceptualization") & (len(Concept) != 0):
               for concept in Concept:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=concept)
                 i += 1
              # print(Doi,  j,  " , ".join(str(concept) for concept in Concept))
           if (j == "Formal analysis") & (len(Analysis) != 0):
               for analysis in Analysis:
                ws.cell(row=i, column=1, value=Doi)
                ws.cell(row=i, column=2, value=j)
                ws.cell(row=i, column=3, value=analysis)
                i += 1
              # print(Doi,  j,  " , ".join(str(analysis) for analysis in Analysis))
           if (j == "Investigation") & (len(Investigation) != 0):
               for invest in Investigation:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=invest)
                 i += 1
              # print(Doi,  j,  " , ".join(str(invest) for invest in Investigation))
           if (j == "Methodology") & (len(Method) != 0):
               for method in Method:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=method)
                 i += 1
              # print(Doi,  j,  " , ".join(str(method) for method in Method))
           if (j == "Visualization") & (len(Visualization) != 0):
               for visual in Visualization:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=visual)
                 i += 1
              # print(Doi,  j,  " , ".join(str(visual) for visual in Visualization))
           if (j == "Writing – original draft") & (len(Draft) != 0):
               for draft in Draft:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=draft)
                 i += 1
              # print(Doi,  j,  " , ".join(str(draft) for draft in Draft))
           if (j == "Writing – review & editing") & (len(Review) != 0):
               for review in Review:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=review)
                 i += 1
              # print(Doi,  j,  " , ".join(str(review) for review in Review))
           if (j == "Resources") & (len(Resources) != 0):
               for resource in Resources:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=resource)
                 i += 1
              # print(Doi,  j,  " , ".join(str(resource) for resource in Resources))
           if (j == "Supervision") & (len(Supervision) != 0):
               for super in Supervision:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=super)
                 i += 1
              # print(Doi,  j,  " , ".join(str(super) for super in Supervision))
           if (j == "Funding acquisition") & (len(Funding) != 0):
               for fund in Funding:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=fund)
                 i += 1
              # print(Doi,  j,  " , ".join(str(fund) for fund in Funding))
           if (j == "Data curation") & (len(Data) != 0):
               for data in Data:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=data)
                 i += 1
              # print(Doi,  j,  " , ".join(str(data) for data in Data))
           if (j == "Project administration") & (len(Project) != 0):
               for project in Project:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=project)
                 i += 1
              # print(Doi,  j,  " , ".join(str(project) for project in Project))
           if (j == "Validation") & (len(Validation) != 0):
               for valid in Validation:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=valid)
                 i += 1
              # print(Doi,  j,  " , ".join(str(valid) for valid in Validation))
           if (j == "Software") & (len(Software) != 0):
               for soft in Software:
                 ws.cell(row=i, column=1, value=Doi)
                 ws.cell(row=i, column=2, value=j)
                 ws.cell(row=i, column=3, value=soft)
                 i += 1
              # print(Doi,  j,  " , ".join(str(soft) for soft in Software))
         # ws.cell(row=i, column=1, value=Doi)
         # ws.cell(row=i, column=2, value=contributions)

wb.save('D:/2022bishe/pone-1.xlsx')





